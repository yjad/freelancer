import os
# import openpyxl
from openpyxl import Workbook, load_workbook


DATA_FOLDER=r".\data"
MAX_ROW_COUNT = 100
MAX_COL_COUNT = 100

wb_master = Workbook()
ws_master = wb_master.active

for folder, subs, files in os.walk(DATA_FOLDER):
    # print (folder, subs, files)

    for f in files:
        print (os.path.join(folder, f))
        
        xl_wb = load_workbook(os.path.join(os.path.join(folder, f)))
        # xl_ws = xl_wb.active
        sheet = xl_wb[xl_wb.sheetnames[0]]  # use first sheet 
        for row in range (1,MAX_ROW_COUNT):
            line=[]
            for col in range(1,MAX_COL_COUNT):
                v = sheet.cell(row=row, column=col).value
                # print (f, row, col, v)
                if v:
                    line.append(v)
                else:
                    break
            # print (line)
            if len(line) > 0:
                ws_master.append(line)
    wb_master.save("master.xlsx")
    