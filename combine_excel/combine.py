import os
# import openpyxl
from openpyxl import Workbook, load_workbook


DATA_FOLDER=r".\data"

wb_master = Workbook()
ws_master = wb_master.active

for folder, subs, files in os.walk(DATA_FOLDER):
    for f in files:
        print (os.path.join(folder, f))
        
        xl_wb = load_workbook(os.path.join(os.path.join(folder, f)))
        for sheet in xl_wb:  # use first sheet 
            for row in sheet.rows:
                line=[f , sheet.title]  # add 2 columns for the file and sheet name
                for cell in row:
                    line.append(cell.value)
                ws_master.append(line)
    wb_master.save("master.xlsx")
    