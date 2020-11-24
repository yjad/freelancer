import json
from openpyxl import Workbook, load_workbook
MAX_ROW_COUNT = 100
MAX_COL_COUNT = 100

def excel_to_json(excel_file_name):
    xl_wb = load_workbook(excel_file_name)
        # xl_ws = xl_wb.active
    sheet = xl_wb[xl_wb.sheetnames[0]]  # use first sheet 
    dict_array = []
    for row in range (1,MAX_ROW_COUNT):
        line=[]
        for col in range(1,MAX_COL_COUNT):
            v = sheet.cell(row=row, column=col).value
            # print (f, row, col, v)
            if v:
                line.append(v)
            else:
                break
        # print (line)
        if row == 1 :
            header = line
        elif len(line) > 0:
            dictx = list_to_json(header, line)
            # print (dictx)
            dict_array.append(dictx)
    print (json.dumps(dict_array, indent=4))
            
           
def list_to_json(header, line):

    dictx = {}
    for i, h in enumerate(header):
        dictx [h] = line[i]
    return dictx
    
    
excel_to_json(r".\data\test.xlsx")

    
